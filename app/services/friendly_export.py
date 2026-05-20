"""
Exportação amigável da base integrada DMS × Censo Escolar.

Funções públicas
----------------
build_friendly_df(df)  — DataFrame com nomes em português, sem colunas internas,
                         e coluna "Média Qtd. DMS por CNPJ" ao lado das matrículas.
export_xlsx_bytes(df)  — BytesIO multi-abas com formatação openpyxl (R$, %, datas).
"""
from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── mapeamento técnico → português de negócio
COLUMN_RENAME: dict[str, str] = {
    "dms__NMRAZAOSOCIAL":                     "Razão Social",
    "dms__NUCNPJ":                            "CNPJ (DMS)",
    "dms__NUCGA":                             "Nº CGA",
    "dms__DTCOMPETENCIA":                     "Competência",
    "dms__DTUPLOAD":                          "Data Upload",
    "dms__TIPO":                              "Tipo DMS",
    "dms__SITUACAO":                          "Situação DMS",
    "dms__QUANTIDADE":                        "Qtd. Beneficiários (DMS)",
    "dms__CDCTISS":                           "Cód. CTISS",
    "dms__VLMENSALIDADE":                     "Mensalidade (R$)",
    "dms__VLBENEFICIO":                       "Benefício (R$)",
    "dms__VLBASECALCULO":                     "Base de Cálculo ISS (R$)",
    "dms__NUALIQUOTA":                        "Alíquota ISS (%)",
    "dms__VLIMPOSTO":                         "ISS (R$)",
    "dms__CDBENEFICIO":                       "Cód. Benefício",
    "dms__DSBENEFICIO":                       "Desc. Benefício",
    "dms__DSSITUACAOFISCAL":                  "Situação Fiscal",
    "censo__CO_ENTIDADE":                     "Cód. Escola (Censo)",
    "censo__NO_ENTIDADE":                     "Nome da Escola",
    "censo__dependencia_administrativa":      "Dep. Administrativa",
    "censo__CNPJ":                            "CNPJ Escola (Censo)",
    "censo__SG_UF":                           "UF",
    "censo__NO_MUNICIPIO":                    "Município",
    "censo__QT_MAT_BAS":                      "Matrículas (Censo)",
    "censo__matriculas":                      "Matrículas Consolidadas",
    "match_status_principal":                 "Status do Vínculo",
    "merge_confianca":                        "Confiança do Vínculo",
    "similaridade_score":                     "Score Similaridade",
    "iss_por_matricula":                      "ISS por Matrícula (R$)",
    "mensalidade_por_aluno":                  "Mensalidade por Aluno (R$)",
    "base_calculo_por_aluno":                 "Base de Cálculo por Aluno (R$)",
}

_COLUMNS_TO_DROP = {
    "dms____cnpj_norm_dms",
    "dms____cnpj_raiz",
    "dms__merge_linha_ordem",
    "censo____cnpj_norm_censo",
    "censo____cnpj_raiz",
    "censo__censo_fonte_escola",
    "censo__censo_fonte_matricula",
    "censo__censo_exercicio",
    "censo__censo_ctx_UF",
    "censo__censo_ctx_municipio_codigo",
    "censo__censo_ctx_municipio_rotulo_ui",
    "censo__censo_ctx_filtro_municipal_aplicado",
    "censo__CO_MUNICIPIO",
    "merge_metodo_primario",
    "cnpj_censo_candidatos_mesmo_numero",
    "censo_escolas_duplicate_count_para_chave",
    "dms_texto_normalizado",
    "censo_texto_normalizado_match",
    "censo_indice_original",
    "__cnpj_raiz",
}

_STATUS_PT: dict[str, str] = {
    "match_cnpj_exato":             "Vinculado — CNPJ exato",
    "multiplas_escolas_mesmo_cnpj": "Múltiplas escolas / mesmo CNPJ",
    "sem_correspondencia_cnpj":     "Sem correspondência no Censo",
    "sem_cnpj_dms":                 "DMS sem CNPJ normalizável",
    "cnpj_dms_invalido":            "CNPJ DMS inválido",
    "match_texto_complementar":     "Vinculado — texto (fuzzy)",
    "sem_correspondencia_texto":    "Sem correspondência — texto",
}

_CONF_PT: dict[str, str] = {
    "alta_conf_cnpj_exato":         "Alta — CNPJ exato",
    "multiplicidade_cnpj_no_censo": "Média — múltiplas escolas",
    "sem_chave_usavel":             "Sem chave",
    "baixa_conf_texto":             "Baixa — texto",
}

# Ordem de exibição das colunas amigáveis
_COL_ORDER = [
    "Razão Social", "CNPJ (DMS)", "Nº CGA", "Situação Fiscal",
    "Competência", "Tipo DMS", "Situação DMS",
    "Qtd. Beneficiários (DMS)", "Mensalidade (R$)", "Benefício (R$)",
    "Base de Cálculo ISS (R$)", "Alíquota ISS (%)", "ISS (R$)",
    "Cód. Benefício", "Desc. Benefício",
    "Status do Vínculo", "Confiança do Vínculo",
    "Nome da Escola", "Cód. Escola (Censo)", "CNPJ Escola (Censo)",
    "Dep. Administrativa", "UF", "Município",
    "Matrículas (Censo)", "Média Qtd. DMS por CNPJ", "Matrículas Consolidadas",
    "ISS por Matrícula (R$)", "Mensalidade por Aluno (R$)", "Base de Cálculo por Aluno (R$)",
    "Score Similaridade",
    "Cód. CTISS", "Data Upload",
]

_FMT: dict[str, str] = {
    "Mensalidade (R$)":               'R$ #,##0.00',
    "Benefício (R$)":                 'R$ #,##0.00',
    "Base de Cálculo ISS (R$)":       'R$ #,##0.00',
    "ISS (R$)":                       'R$ #,##0.00',
    "ISS por Matrícula (R$)":         'R$ #,##0.00',
    "Mensalidade por Aluno (R$)":     'R$ #,##0.00',
    "Base de Cálculo por Aluno (R$)": 'R$ #,##0.00',
    "Alíquota ISS (%)":               '0.00"%"',
    "Competência":                    'DD/MM/YYYY',
    "Data Upload":                    'DD/MM/YYYY HH:MM',
    "Matrículas (Censo)":             '#,##0',
    "Matrículas Consolidadas":        '#,##0',
    "Qtd. Beneficiários (DMS)":       '#,##0',
    "Média Qtd. DMS por CNPJ":        '#,##0.0',
    "Score Similaridade":             '0.00',
}

_HEADER_FILLS: dict[str, str] = {
    "Todas as linhas":      "1F4E79",
    "Vinculados":           "1E7145",
    "Sem Correspondência":  "843C0C",
    "Resumo por Empresa":   "44546A",
}

_VINCULADOS_STATUS = {
    "match_cnpj_exato",
    "multiplas_escolas_mesmo_cnpj",
    "match_texto_complementar",
}
_SEM_CORRESP_STATUS = {
    "sem_correspondencia_cnpj",
    "sem_cnpj_dms",
    "cnpj_dms_invalido",
    "sem_correspondencia_texto",
}


# ─── público ──────────────────────────────────────────────────────────────────

def build_friendly_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara o DataFrame para exportação amigável:
    - Remove colunas técnicas internas
    - Adiciona "Média Qtd. DMS por CNPJ" ao lado das matrículas do Censo
    - Traduz status/confiança para português
    - Renomeia colunas e reordena conforme _COL_ORDER
    """
    out = df.copy()

    out = out.drop(columns=[c for c in _COLUMNS_TO_DROP if c in out.columns])

    if "dms__NUCNPJ" in out.columns and "dms__QUANTIDADE" in out.columns:
        qty_num = pd.to_numeric(out["dms__QUANTIDADE"], errors="coerce")
        out["__qty_num_tmp"] = qty_num
        media = out.groupby("dms__NUCNPJ", sort=False)["__qty_num_tmp"].transform("mean")
        out = out.drop(columns=["__qty_num_tmp"])
        insert_at = list(out.columns).index("dms__QUANTIDADE") + 1
        out.insert(insert_at, "Média Qtd. DMS por CNPJ", media)

    if "match_status_principal" in out.columns:
        out["match_status_principal"] = (
            out["match_status_principal"]
            .map(lambda v: _STATUS_PT.get(str(v), str(v)) if pd.notna(v) else v)
        )
    if "merge_confianca" in out.columns:
        out["merge_confianca"] = (
            out["merge_confianca"]
            .map(lambda v: _CONF_PT.get(str(v), str(v)) if pd.notna(v) else v)
        )

    out = out.rename(columns=COLUMN_RENAME)

    # Coerce colunas numéricas — o pipeline pode produzir dtype string[pyarrow]
    _NUMERIC_FRIENDLY = {
        "Qtd. Beneficiários (DMS)", "Mensalidade (R$)", "Benefício (R$)",
        "Base de Cálculo ISS (R$)", "Alíquota ISS (%)", "ISS (R$)",
        "Matrículas (Censo)", "Matrículas Consolidadas",
        "ISS por Matrícula (R$)", "Mensalidade por Aluno (R$)",
        "Base de Cálculo por Aluno (R$)", "Score Similaridade",
    }
    for col in _NUMERIC_FRIENDLY:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")

    present_ordered = [c for c in _COL_ORDER if c in out.columns]
    remaining = [c for c in out.columns if c not in set(_COL_ORDER)]
    return out[present_ordered + remaining]


def export_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """
    Gera bytes de .xlsx multi-abas amigável a partir da base integrada.
    Pronto para passar a st.download_button(data=...).
    """
    df_friendly = build_friendly_df(df)

    status_col = "Status do Vínculo"

    def _mask(raw_statuses: set[str]) -> pd.Series:
        if status_col not in df_friendly.columns:
            return pd.Series([True] * len(df_friendly), index=df_friendly.index)
        pt = {_STATUS_PT.get(s, s) for s in raw_statuses}
        return df_friendly[status_col].isin(pt)

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Todas as linhas",     df_friendly),
        ("Vinculados",          df_friendly[_mask(_VINCULADOS_STATUS)].copy()),
        ("Sem Correspondência", df_friendly[_mask(_SEM_CORRESP_STATUS)].copy()),
        ("Resumo por Empresa",  _build_resumo(df_friendly)),
    ]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, sdf in sheets:
            if not sdf.empty:
                sdf.to_excel(writer, sheet_name=name, index=False)

    buf.seek(0)
    wb = load_workbook(buf)
    for name, fill_hex in _HEADER_FILLS.items():
        if name in wb.sheetnames:
            _format_sheet(wb[name], fill_hex)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── internos ─────────────────────────────────────────────────────────────────

def _build_resumo(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = [c for c in ["Razão Social", "CNPJ (DMS)"] if c in df.columns]
    if not group_cols:
        return pd.DataFrame()

    sum_cols   = ["Base de Cálculo ISS (R$)", "ISS (R$)", "Mensalidade (R$)", "Benefício (R$)"]
    mean_cols  = ["Qtd. Beneficiários (DMS)", "ISS por Matrícula (R$)", "Mensalidade por Aluno (R$)"]
    first_cols = ["Nome da Escola", "Status do Vínculo", "Matrículas (Censo)"]

    work = df.copy()
    for c in sum_cols + mean_cols:
        if c in work.columns:
            work[c] = pd.to_numeric(work[c], errors="coerce")

    agg: dict[str, Any] = {}
    for c in sum_cols:
        if c in work.columns:
            agg[c] = "sum"
    for c in mean_cols:
        if c in work.columns:
            agg[c] = "mean"
    for c in first_cols:
        if c in work.columns:
            agg[c] = "first"
    if "Competência" in work.columns:
        agg["Competência"] = "count"

    if not agg:
        return pd.DataFrame()

    resumo = work.groupby(group_cols, as_index=False, dropna=False).agg(agg)

    if "Competência" in resumo.columns:
        resumo = resumo.rename(columns={"Competência": "Meses declarados"})

    sort_col = "Base de Cálculo ISS (R$)"
    if sort_col in resumo.columns:
        resumo = resumo.sort_values(sort_col, ascending=False)

    return resumo.reset_index(drop=True)


def _format_sheet(ws: Any, fill_hex: str) -> None:
    header_fill = PatternFill("solid", fgColor=fill_hex)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.freeze_panes = "A2"

    col_names = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]
    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)
        fmt = _FMT.get(str(col_name))
        if fmt:
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.number_format = fmt
        ws.column_dimensions[col_letter].width = min(max(len(str(col_name)), 10) + 2, 42)
