"""
Relatório Operacional de Divergências — DMS Educação × Censo.

Produto institucional da Inteligência Fiscal / GEFIS.
Exporta UMA aba com UMA linha por CNPJ declarante na DMS.

Função pública
--------------
exportar_divergencias_operacionais(df, exercicio, col_censo_entidade,
                                   label_censo, cnpjs_mistos) -> bytes (.xlsx)
"""
from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── constantes visuais
_BLUE_HEADER   = "1F4E79"   # azul institucional — cabeçalho
_ZEBRA_EVEN    = "EBF3FB"   # azul muito claro — linhas pares
_HIGHLIGHT_TOP = "FCE4D6"   # laranja suave — maiores divergências (top 20%)
_AMBER_MISTO   = "FFF2CC"   # amarelo âmbar — instituição mista Básica+Superior

_FMT_INTEIRO = "#,##0"
_FMT_DECIMAL = "#,##0.0"


# ─── público ──────────────────────────────────────────────────────────────────

def exportar_divergencias_operacionais(
    df: pd.DataFrame,
    exercicio: int | None = None,
    col_censo_entidade: str = "censo__CO_ENTIDADE",
    label_censo: str = "Qtd Matrículas Censo Escolar",
    cnpjs_mistos: frozenset[str] | set[str] | None = None,
) -> bytes:
    """
    Gera o Relatório Operacional de Divergências em .xlsx.

    Parâmetros
    ----------
    df               : base integrada DMS × Censo (saída do pipeline de merge).
    exercicio        : ano de referência para filtro de maio. Auto-detectado se None.
    col_censo_entidade : coluna que identifica a entidade de ensino no Censo.
                        "censo__CO_ENTIDADE" para Ed. Básica (padrão);
                        "censo__CO_IES" para Ensino Superior.
    label_censo      : rótulo da coluna de matrículas do Censo no relatório.
                        "Qtd Matrículas Censo Escolar" (padrão — Ed. Básica);
                        "Qtd Matrículas Censo Superior" (Ensino Superior).
    cnpjs_mistos     : conjunto de CNPJs (14 dígitos, zero-fill) que também
                        aparecem no outro nível de ensino. Esses CNPJs recebem
                        alerta "Sim" na coluna "Instituição Mista".

    Retorna bytes prontos para st.download_button(data=...).
    """
    work = df.copy()

    # ── 1. Garantir tipos numéricos
    work["__qtd"]       = pd.to_numeric(work.get("dms__QUANTIDADE",    pd.Series(dtype=float)), errors="coerce")
    work["__censo_mat"] = pd.to_numeric(work.get("censo__QT_MAT_BAS",  pd.Series(dtype=float)), errors="coerce")
    work["__comp"]      = pd.to_datetime(work.get("dms__DTCOMPETENCIA", pd.Series(dtype=object)), errors="coerce")

    # ── 2. Exercício (ano de referência)
    if exercicio is None:
        anos = work["__comp"].dt.year.dropna()
        exercicio = int(anos.mode().iloc[0]) if not anos.empty else 2025

    # ── 3. Qtd Matrículas DMS — mês de maio do exercício
    mask_maio = (work["__comp"].dt.month == 5) & (work["__comp"].dt.year == exercicio)
    qtd_maio: pd.Series = (
        work.loc[mask_maio]
        .groupby("dms__NUCNPJ", sort=False)["__qtd"]
        .sum()
        .rename("qtd_maio")
    )

    # ── 4. Matrículas Censo por CNPJ — soma de entidades únicas vinculadas
    if col_censo_entidade in work.columns:
        entidades_unicas = (
            work.loc[work[col_censo_entidade].notna()]
            .drop_duplicates(["dms__NUCNPJ", col_censo_entidade])
        )
        censo_por_cnpj: pd.Series = (
            entidades_unicas.groupby("dms__NUCNPJ", sort=False)["__censo_mat"]
            .sum()
            .rename("qtd_censo")
        )
    else:
        censo_por_cnpj = pd.Series(dtype=float, name="qtd_censo")

    # ── 5. Razão Social (primeiro valor não nulo por CNPJ)
    razao_social: pd.Series = (
        work.groupby("dms__NUCNPJ", sort=False)["dms__NMRAZAOSOCIAL"]
        .first()
        .rename("razao_social")
    )

    # ── 6. Montar tabela única por CNPJ
    resultado = (
        razao_social.to_frame()
        .join(qtd_maio,       how="left")
        .join(censo_por_cnpj, how="left")
        .reset_index()
        .rename(columns={"dms__NUCNPJ": "cnpj_raw"})
    )

    # ── 7. Normalizar CNPJ para 14 dígitos (zero-fill)
    resultado["CNPJ"] = (
        resultado["cnpj_raw"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .str.zfill(14)
    )

    # ── 8. Diferença absoluta
    resultado["Diferença Censo × Maio"] = (
        resultado["qtd_censo"] - resultado["qtd_maio"]
    ).abs()

    # ── 9. Alerta de instituição mista
    if cnpjs_mistos:
        resultado["Instituição Mista"] = (
            resultado["CNPJ"].isin(cnpjs_mistos)
            .map({True: "Sim", False: ""})
        )
    else:
        resultado["Instituição Mista"] = ""

    # ── 10. Filtro: manter apenas linhas com informação quantitativa
    tem_info = resultado["qtd_maio"].gt(0) | resultado["qtd_censo"].gt(0)
    resultado = resultado.loc[tem_info].copy()

    # ── 11. Ordenar: maiores divergências primeiro
    resultado = resultado.sort_values(
        "Diferença Censo × Maio",
        ascending=False,
        na_position="last",
    )

    # ── 12. Renomear e selecionar colunas finais
    resultado = resultado.rename(columns={
        "razao_social": "Razão Social",
        "qtd_censo":    label_censo,
        "qtd_maio":     "Qtd Matrículas DMS Educação (Maio)",
    })

    cols_finais = [
        "CNPJ",
        "Razão Social",
        label_censo,
        "Qtd Matrículas DMS Educação (Maio)",
        "Diferença Censo × Maio",
        "Instituição Mista",
    ]
    cols_presentes = [c for c in cols_finais if c in resultado.columns]
    resultado = resultado[cols_presentes].reset_index(drop=True)

    col_widths = {
        "CNPJ":                               18,
        "Razão Social":                       50,
        label_censo:                          26,
        "Qtd Matrículas DMS Educação (Maio)": 26,
        "Diferença Censo × Maio":             22,
        "Instituição Mista":                  18,
    }
    numeric_cols = {
        label_censo:                          _FMT_INTEIRO,
        "Qtd Matrículas DMS Educação (Maio)": _FMT_INTEIRO,
        "Diferença Censo × Maio":             _FMT_DECIMAL,
    }

    return _escrever_xlsx(resultado, col_widths, numeric_cols)


# ─── internos ─────────────────────────────────────────────────────────────────

def _escrever_xlsx(
    df: pd.DataFrame,
    col_widths: dict[str, int],
    numeric_cols: dict[str, str],
) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, sheet_name="Divergências", engine="openpyxl")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb["Divergências"]

    col_names = [ws.cell(1, c).value or "" for c in range(1, ws.max_column + 1)]

    _formatar_cabecalho(ws, col_names)
    _formatar_dados(ws, col_names, numeric_cols)
    _ajustar_larguras(ws, col_names, col_widths)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _formatar_cabecalho(ws: Any, col_names: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor=_BLUE_HEADER)
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 32


def _formatar_dados(
    ws: Any,
    col_names: list[str],
    numeric_cols: dict[str, str],
) -> None:
    zebra    = PatternFill("solid", fgColor=_ZEBRA_EVEN)
    destaque = PatternFill("solid", fgColor=_HIGHLIGHT_TOP)
    misto    = PatternFill("solid", fgColor=_AMBER_MISTO)

    # Índices das colunas de interesse
    dif_col_idx: int | None = None
    try:
        dif_col_idx = col_names.index("Diferença Censo × Maio") + 1
    except ValueError:
        pass

    misto_col_idx: int | None = None
    try:
        misto_col_idx = col_names.index("Instituição Mista") + 1
    except ValueError:
        pass

    # Limiar top-20% de divergência
    threshold: float | None = None
    if dif_col_idx is not None:
        valores = [
            ws.cell(r, dif_col_idx).value
            for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(r, dif_col_idx).value, (int, float))
               and ws.cell(r, dif_col_idx).value > 0
        ]
        if valores:
            top_n = max(1, len(valores) // 5)
            threshold = sorted(valores, reverse=True)[top_n - 1]

    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        dif_val   = ws.cell(row_idx, dif_col_idx).value   if dif_col_idx   else None
        misto_val = ws.cell(row_idx, misto_col_idx).value if misto_col_idx else None

        eh_destaque = (
            threshold is not None
            and isinstance(dif_val, (int, float))
            and dif_val >= threshold
        )
        eh_misto = misto_val == "Sim"

        # Prioridade: divergência alta > instituição mista > zebra
        if eh_destaque:
            row_fill = destaque
        elif eh_misto:
            row_fill = misto
        elif row_idx % 2 == 0:
            row_fill = zebra
        else:
            row_fill = None

        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row_idx, col_idx)

            if row_fill:
                cell.fill = row_fill

            if col_name in numeric_cols:
                cell.number_format = numeric_cols[col_name]
                cell.alignment = center_align
            elif col_name in ("CNPJ", "Instituição Mista"):
                cell.alignment = center_align
            else:
                cell.alignment = left_align


def _ajustar_larguras(
    ws: Any,
    col_names: list[str],
    col_widths: dict[str, int],
) -> None:
    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(str(col_name), 20)
